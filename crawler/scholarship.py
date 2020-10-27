# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 16:10:30 2020

@author: User
獎學金爬蟲加整理加放入DATABASE
"""

import openpyxl
import pymongo
import pandas as pd
import urllib.request as req #載入模組並設定別名
import bs4 # 爬蟲分析


#-------------------------------------------------------------------爬蟲(取得.excel)--------------------
res = pd.read_html("https://itouch.cycu.edu.tw/active_system/query_data/student/ssgogo.jsp")
data = res[1].iloc[:,0:10]
data.to_excel("scholarship.xlsx")

#-------------------------------------------------------------------爬蟲(取得.網址)--------------------
url="https://itouch.cycu.edu.tw/active_system/query_data/student/ssgogo.jsp"

#!!!必須建立一個Request的物件,附加Request Headers的資訊,讓Pprogram看起來像正常的使用者
request=req.Request(url,headers={
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36"
})

with req.urlopen(request) as response:
    data = response.read().decode("utf-8")
    

root = bs4.BeautifulSoup(data, "html.parser")
titles = root.find_all("a")

nameList = []
hrefList = []

for title in titles :
    nameList.append(title.string)
    hrefList.append(title.get('href')) # 放入list存起來
    
list_href = []

for i in hrefList :
  if not ( "mailto" in i or "javascript" in i ) :
    list_href.append(i)
 
#-------------------------------------------------------------------讀檔+整理資料--------------------

workbook = openpyxl.load_workbook('scholarship.xlsx') # 讀入 excel 檔

# 獲得所有sheet的名稱
#print("所有sheet的名稱:")
#print( workbook.get_sheet_names())
# 根據sheet名字獲得sheet
worksheet = workbook.get_sheet_by_name('Sheet1')
# 查看當前worksheet指定的此sheet名
#print(worksheet.title)

scholarship_list = []
#print(worksheet.max_row)
for i in range( 4, worksheet.max_row ) :

  scholarship_dict = { "名稱": worksheet.cell( row = i, column = 3 ).value ,
                       "網址": list_href.pop(0) ,
                       "截止日期": worksheet.cell( row = i, column = 4).value ,
                       "金額": worksheet.cell( row = i, column = 5).value ,
                       "學業成績" : worksheet.cell( row = i, column = 6).value ,
                       "操行成績" : worksheet.cell( row = i, column = 7).value ,
                       "體育成績" : worksheet.cell( row = i, column = 8).value ,
                       "申請身分" : worksheet.cell( row = i, column = 9).value ,
                       "申請資格" : worksheet.cell( row = i, column = 10).value ,      
                       "應繳文件" : worksheet.cell( row = i, column = 11).value }
                       
                                            
  #整理金額
  if scholarship_dict["金額"] == None :
      scholarship_dict["金額"] = "面議"
      
  # 整理所有分數 把no改成0若不是no則改型別為int
  if scholarship_dict["學業成績"] != "no" :
      scholarship_dict["學業成績"] = int(scholarship_dict["學業成績"])
  else :
      scholarship_dict["學業成績"] = 0 
      
  if scholarship_dict["操行成績"] != "no" :
      scholarship_dict["操行成績"] = int(scholarship_dict["操行成績"])
  else :
      scholarship_dict["操行成績"] = 0 
      
  if scholarship_dict["體育成績"] != "no" :
      scholarship_dict["體育成績"] = int(scholarship_dict["體育成績"])
  else :
      scholarship_dict["體育成績"] = 0 
      
  scholarship_list.append(scholarship_dict)
#print("獎學金筆數:" + str(len(scholarship_list)))

#-------------------------------------------------------------------放入DataBase--------------------

client = pymongo.MongoClient("mongodb+srv://Jerry_Chang:jerry123@cluster0.mmp88.mongodb.net/Jerry_Chang?retryWrites=true&w=majority")

db = client.Total_Scholarship
db.不拘.delete_many( {} )
db.大學部.delete_many( {} )
db.研究所.delete_many( {} )
db.一般類.delete_many( {} )
db.清寒類.delete_many( {} )
db.大一新生.delete_many( {} )
db.所有類.delete_many( {} )

for temp in scholarship_list :
    
    db.所有類.insert_one( temp )
    
    if "不拘" in temp["申請身分"] :
        db.不拘.insert_one( temp )
    if "大學部" in temp["申請身分"] :
        db.大學部.insert_one( temp )
    if "清寒類" in temp["申請身分"] :
        db.清寒類.insert_one( temp )
    if "一般類" in temp["申請身分"] :
        db.一般類.insert_one( temp )
    if "研究所" in temp["申請身分"] :
        db.研究所.insert_one( temp )
    if "大一新生" in temp["申請身分"] :
        db.大一新生.insert_one( temp )
        