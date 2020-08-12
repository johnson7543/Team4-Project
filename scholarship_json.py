# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 16:10:30 2020

@author: User
"""

import openpyxl
import json
import pymongo

workbook = openpyxl.load_workbook('scholarship.xlsx') # 讀入 excel 檔

worksheet = workbook.get_sheet_by_name('Sheet1')

row1 = worksheet.cell(row=2,column=3).value

scholarship_json = { "Item_name": "jim" , "End_date": 25, "Money":"Taiwan", "Grade" : 80, "Sport_Grade" : 80, "Apply" : "blahblahblah" }

list = []

for i in range( 4, 36) :

  scholarship_json = { "Item_name": worksheet.cell( row = i, column = 3).value ,
                       "End_date": worksheet.cell( row = i, column = 4).value ,
                       "Money": worksheet.cell( row = i, column = 5).value ,
                       "Grade" : worksheet.cell( row = i, column = 6).value ,
                       "Sport_Grade" : worksheet.cell( row = i, column = 8).value,
                       "Apply" : worksheet.cell( row = i, column = 10).value }
  
  list.append(scholarship_json)

# print( list )


client = pymongo.MongoClient("mongodb+srv://Jerry_Chang:jerry123@cluster0.mmp88.mongodb.net/Jerry_Chang?retryWrites=true&w=majority")

db = client.blog
db.posts.delete_many( {} )

for temp in  list :

  db.posts.insert_one( temp ) 


for post in client.blog.posts.find():
  print(post) 

# for item in list(worksheet.rows)[1] :
  
""" row3 = [item.value for item in list(worksheet.rows)[2] ]

print('第3行值',row3)

col3 = [item.value for item in list(worksheet.columns)[2]]

print('第3行值',col3)

cell_2_3=worksheet.cell(row=2,column=3).value

print('第2行第3列值',cell_2_3)

max_row=worksheet.max_row

print('最大行',max_row)
 """




